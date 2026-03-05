from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generar_reporte_excel(datos):
    """
    Genera archivo Excel con 2 hojas:
    - Hoja 1: Resumen (empleado, dias trabajados, retardos, faltas)
    - Hoja 2: Detalle de registros de asistencia
    """
    wb = openpyxl.Workbook()

    # Estilos comunes
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    retardo_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    falta_fill = PatternFill(start_color='FECACA', end_color='FECACA', fill_type='solid')

    # === HOJA 1: Resumen ===
    ws1 = wb.active
    ws1.title = "Resumen"

    # Titulo
    ws1.merge_cells('A1:G1')
    titulo_cell = ws1['A1']
    titulo_cell.value = f"Reporte de Asistencia - {datos['fecha_inicio']} al {datos['fecha_fin']}"
    titulo_cell.font = Font(bold=True, size=14, color='1E3A5F')
    titulo_cell.alignment = Alignment(horizontal='center')

    # Info general
    ws1['A2'] = f"Dias laborales: {datos['dias_laborales']}"
    ws1['A2'].font = Font(size=10, italic=True)
    ws1['D2'] = f"Total empleados: {datos['total_empleados']}"
    ws1['D2'].font = Font(size=10, italic=True)

    # Headers de tabla
    headers = ['Codigo', 'Nombre', 'Departamento', 'Dias Trabajados', 'Retardos', 'Faltas', 'Horas Totales']
    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Datos
    for row_idx, emp in enumerate(datos['datos_empleados'], 5):
        ws1.cell(row=row_idx, column=1, value=emp['codigo']).border = thin_border
        ws1.cell(row=row_idx, column=2, value=emp['nombre']).border = thin_border
        ws1.cell(row=row_idx, column=3, value=emp['departamento']).border = thin_border

        cell_dias = ws1.cell(row=row_idx, column=4, value=emp['dias_trabajados'])
        cell_dias.border = thin_border
        cell_dias.alignment = Alignment(horizontal='center')

        cell_ret = ws1.cell(row=row_idx, column=5, value=emp['retardos'])
        cell_ret.border = thin_border
        cell_ret.alignment = Alignment(horizontal='center')
        if emp['retardos'] > 0:
            cell_ret.fill = retardo_fill

        cell_fal = ws1.cell(row=row_idx, column=6, value=emp['faltas'])
        cell_fal.border = thin_border
        cell_fal.alignment = Alignment(horizontal='center')
        if emp['faltas'] > 0:
            cell_fal.fill = falta_fill

        cell_hrs = ws1.cell(row=row_idx, column=7, value=emp['horas_totales'])
        cell_hrs.border = thin_border
        cell_hrs.alignment = Alignment(horizontal='center')
        cell_hrs.number_format = '0.00'

    # Ajustar anchos
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 10
    ws1.column_dimensions['G'].width = 15

    # === HOJA 2: Detalle ===
    ws2 = wb.create_sheet("Detalle de Asistencia")

    # Titulo
    ws2.merge_cells('A1:H1')
    titulo2 = ws2['A1']
    titulo2.value = f"Detalle de Registros - {datos['fecha_inicio']} al {datos['fecha_fin']}"
    titulo2.font = Font(bold=True, size=14, color='1E3A5F')
    titulo2.alignment = Alignment(horizontal='center')

    # Headers
    headers2 = ['Fecha', 'Codigo', 'Nombre', 'Hora Entrada', 'Hora Salida', 'Horas Trabajadas', 'Retardo', 'Incidencia']
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Recopilar todos los registros ordenados por fecha
    todos_registros = []
    for emp in datos['datos_empleados']:
        for reg in emp['registros']:
            todos_registros.append({
                'fecha': reg.fecha,
                'codigo': emp['codigo'],
                'nombre': emp['nombre'],
                'hora_entrada': reg.hora_entrada,
                'hora_salida': reg.hora_salida,
                'horas_trabajadas': reg.horas_trabajadas or 0,
                'retardo': reg.retardo,
                'incidencia': reg.get_incidencia_display() if reg.incidencia != 'ninguna' else '',
            })

    todos_registros.sort(key=lambda x: (x['fecha'], x['codigo']))

    for row_idx, reg in enumerate(todos_registros, 4):
        ws2.cell(row=row_idx, column=1, value=str(reg['fecha'])).border = thin_border
        ws2.cell(row=row_idx, column=2, value=reg['codigo']).border = thin_border
        ws2.cell(row=row_idx, column=3, value=reg['nombre']).border = thin_border

        cell_ent = ws2.cell(row=row_idx, column=4, value=str(reg['hora_entrada']) if reg['hora_entrada'] else '')
        cell_ent.border = thin_border
        cell_ent.alignment = Alignment(horizontal='center')

        cell_sal = ws2.cell(row=row_idx, column=5, value=str(reg['hora_salida']) if reg['hora_salida'] else '')
        cell_sal.border = thin_border
        cell_sal.alignment = Alignment(horizontal='center')

        cell_hrs = ws2.cell(row=row_idx, column=6, value=round(reg['horas_trabajadas'], 2))
        cell_hrs.border = thin_border
        cell_hrs.alignment = Alignment(horizontal='center')
        cell_hrs.number_format = '0.00'

        cell_ret = ws2.cell(row=row_idx, column=7, value='Si' if reg['retardo'] else '')
        cell_ret.border = thin_border
        cell_ret.alignment = Alignment(horizontal='center')
        if reg['retardo']:
            cell_ret.fill = retardo_fill

        cell_inc = ws2.cell(row=row_idx, column=8, value=reg['incidencia'])
        cell_inc.border = thin_border

    # Ajustar anchos
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 30
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 16
    ws2.column_dimensions['G'].width = 10
    ws2.column_dimensions['H'].width = 20

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _estilos_base():
    """Retorna estilos comunes reutilizables."""
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    return header_font, header_fill, header_alignment, thin_border


def _escribir_fila_header(ws, row, headers, header_font, header_fill, header_alignment, thin_border):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def generar_reporte_inventario_excel(datos):
    """
    Excel de inventario con 2 hojas:
    - Hoja 1: Resumen por estado + equipos con mantenimiento pendiente
    - Hoja 2: Últimos mantenimientos registrados (30 días)
    """
    wb = openpyxl.Workbook()
    header_font, header_fill, header_alignment, thin_border = _estilos_base()
    alerta_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    advertencia_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')

    # === HOJA 1: Resumen ===
    ws1 = wb.active
    ws1.title = "Inventario"

    ws1.merge_cells('A1:G1')
    titulo = ws1['A1']
    titulo.value = f"Inventario de Equipos - {datos['fecha_reporte']}"
    titulo.font = Font(bold=True, size=14, color='1E3A5F')
    titulo.alignment = Alignment(horizontal='center')

    # Stat boxes de estado
    ws1['A3'] = f"Total equipos: {datos['total_equipos']}"
    ws1['A3'].font = Font(bold=True, size=11)
    ws1['C3'] = f"Activos: {datos['por_estado'].get('activo', 0)}"
    ws1['E3'] = f"En mantenimiento: {datos['por_estado'].get('mantenimiento', 0)}"
    ws1['G3'] = f"Baja: {datos['por_estado'].get('baja', 0)}"

    # Tabla: Equipos con mantenimiento vencido
    ws1['A5'] = "Equipos con Mantenimiento VENCIDO"
    ws1['A5'].font = Font(bold=True, size=12, color='DC2626')

    headers1 = ['No. Serie', 'Marca / Modelo', 'Tipo', 'Usuario / Empleado', 'Último Mant.', 'Próximo Mant.', 'Estado']
    _escribir_fila_header(ws1, 6, headers1, header_font, header_fill, header_alignment, thin_border)

    row = 7
    for eq in datos['equipos_vencidos']:
        usuario = eq.empleado.nombre_completo if eq.empleado else eq.usuario_nombre or '-'
        ws1.cell(row=row, column=1, value=eq.numero_serie).border = thin_border
        ws1.cell(row=row, column=2, value=f"{eq.marca} {eq.modelo}").border = thin_border
        ws1.cell(row=row, column=3, value=eq.get_tipo_display()).border = thin_border
        ws1.cell(row=row, column=4, value=usuario).border = thin_border
        c_ult = ws1.cell(row=row, column=5, value=str(eq.fecha_ultimo_mantenimiento) if eq.fecha_ultimo_mantenimiento else 'Nunca')
        c_ult.border = thin_border
        c_prox = ws1.cell(row=row, column=6, value=str(eq.fecha_proximo_mantenimiento) if eq.fecha_proximo_mantenimiento else '-')
        c_prox.border = thin_border
        c_prox.fill = alerta_fill
        ws1.cell(row=row, column=7, value=eq.get_estado_display()).border = thin_border
        row += 1

    if not datos['equipos_vencidos']:
        ws1.merge_cells(f'A{row}:G{row}')
        ws1.cell(row=row, column=1, value='Sin equipos vencidos').font = Font(italic=True, color='6B7280')
        row += 1

    # Tabla: Equipos que requieren mantenimiento pronto
    row += 1
    ws1.cell(row=row, column=1, value="Equipos que Requieren Mantenimiento Próximo (≤30 días)")
    ws1.cell(row=row, column=1).font = Font(bold=True, size=12, color='D97706')
    row += 1
    _escribir_fila_header(ws1, row, headers1, header_font, header_fill, header_alignment, thin_border)
    row += 1
    for eq in datos['equipos_pronto']:
        usuario = eq.empleado.nombre_completo if eq.empleado else eq.usuario_nombre or '-'
        ws1.cell(row=row, column=1, value=eq.numero_serie).border = thin_border
        ws1.cell(row=row, column=2, value=f"{eq.marca} {eq.modelo}").border = thin_border
        ws1.cell(row=row, column=3, value=eq.get_tipo_display()).border = thin_border
        ws1.cell(row=row, column=4, value=usuario).border = thin_border
        ws1.cell(row=row, column=5, value=str(eq.fecha_ultimo_mantenimiento) if eq.fecha_ultimo_mantenimiento else 'Nunca').border = thin_border
        c_prox = ws1.cell(row=row, column=6, value=str(eq.fecha_proximo_mantenimiento) if eq.fecha_proximo_mantenimiento else '-')
        c_prox.border = thin_border
        c_prox.fill = advertencia_fill
        ws1.cell(row=row, column=7, value=eq.get_estado_display()).border = thin_border
        row += 1

    if not datos['equipos_pronto']:
        ws1.merge_cells(f'A{row}:G{row}')
        ws1.cell(row=row, column=1, value='Sin equipos próximos a mantenimiento').font = Font(italic=True, color='6B7280')

    for col, width in zip('ABCDEFG', [18, 28, 14, 28, 16, 16, 14]):
        ws1.column_dimensions[col].width = width

    # === HOJA 2: Últimos mantenimientos ===
    ws2 = wb.create_sheet("Últimos Mantenimientos")
    ws2.merge_cells('A1:G1')
    ws2['A1'].value = "Mantenimientos Realizados (últimos 30 días)"
    ws2['A1'].font = Font(bold=True, size=14, color='1E3A5F')
    ws2['A1'].alignment = Alignment(horizontal='center')

    headers2 = ['Fecha', 'Equipo', 'Tipo Mant.', 'Técnico', 'Descripción', 'Próximo', 'Costo']
    _escribir_fila_header(ws2, 3, headers2, header_font, header_fill, header_alignment, thin_border)

    for row_idx, mant in enumerate(datos['ultimos_mantenimientos'], 4):
        ws2.cell(row=row_idx, column=1, value=str(mant.fecha_realizado)).border = thin_border
        ws2.cell(row=row_idx, column=2, value=f"{mant.equipo.marca} {mant.equipo.modelo} ({mant.equipo.numero_serie})").border = thin_border
        ws2.cell(row=row_idx, column=3, value=mant.get_tipo_mantenimiento_display()).border = thin_border
        ws2.cell(row=row_idx, column=4, value=mant.tecnico).border = thin_border
        ws2.cell(row=row_idx, column=5, value=mant.descripcion[:80] if mant.descripcion else '').border = thin_border
        ws2.cell(row=row_idx, column=6, value=str(mant.fecha_proximo) if mant.fecha_proximo else '-').border = thin_border
        costo_cell = ws2.cell(row=row_idx, column=7, value=float(mant.costo) if mant.costo else 0)
        costo_cell.border = thin_border
        costo_cell.number_format = '"$"#,##0.00'

    if not datos['ultimos_mantenimientos']:
        ws2.merge_cells('A4:G4')
        ws2.cell(row=4, column=1, value='Sin mantenimientos en los últimos 30 días').font = Font(italic=True, color='6B7280')

    for col, width in zip('ABCDEFG', [12, 35, 18, 20, 40, 14, 12]):
        ws2.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_reporte_tickets_excel(datos):
    """
    Excel de tickets IT con 2 hojas:
    - Hoja 1: Estadísticas (por estado, prioridad, categoría, técnico)
    - Hoja 2: Detalle de tickets
    """
    wb = openpyxl.Workbook()
    header_font, header_fill, header_alignment, thin_border = _estilos_base()
    rojo_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    amarillo_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    verde_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')

    # === HOJA 1: Estadísticas ===
    ws1 = wb.active
    ws1.title = "Estadísticas"

    ws1.merge_cells('A1:F1')
    ws1['A1'].value = f"Reporte Tickets IT - {datos['fecha_inicio']} al {datos['fecha_fin']}"
    ws1['A1'].font = Font(bold=True, size=14, color='1E3A5F')
    ws1['A1'].alignment = Alignment(horizontal='center')

    ws1['A3'] = f"Total tickets: {datos['total_tickets']}"
    ws1['A3'].font = Font(bold=True, size=12)
    ws1['C3'] = f"Concluidos: {datos['tickets_concluidos']}"
    ws1['E3'] = f"Promedio resolución: {datos['promedio_horas_resolucion']} hrs" if datos['promedio_horas_resolucion'] else "Promedio resolución: N/A"

    # Por estado
    row = 5
    ws1.cell(row=row, column=1, value="Por Estado").font = Font(bold=True, size=11, color='1E3A5F')
    row += 1
    _escribir_fila_header(ws1, row, ['Estado', 'Cantidad'], header_font, header_fill, header_alignment, thin_border)
    row += 1
    estados_display = {
        'creado': 'Creado', 'pendiente': 'Pendiente', 'proceso': 'En Proceso',
        'espera': 'En Espera', 'concluido': 'Concluido',
    }
    for estado, cantidad in datos['por_estado'].items():
        c1 = ws1.cell(row=row, column=1, value=estados_display.get(estado, estado))
        c1.border = thin_border
        c2 = ws1.cell(row=row, column=2, value=cantidad)
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='center')
        if estado == 'concluido' and cantidad > 0:
            c2.fill = verde_fill
        row += 1

    # Por prioridad (columna D)
    row = 5
    ws1.cell(row=row, column=4, value="Por Prioridad").font = Font(bold=True, size=11, color='1E3A5F')
    row += 1
    _escribir_fila_header(ws1, row, ['Prioridad', 'Cantidad'], header_font, header_fill, header_alignment, thin_border)
    col_offset = 4
    row += 1
    prioridades_display = {'critica': 'Crítica', 'alta': 'Alta', 'media': 'Media', 'baja': 'Baja', 'sin_asignar': 'Sin asignar'}
    for prioridad, cantidad in datos['por_prioridad'].items():
        c1 = ws1.cell(row=row, column=col_offset, value=prioridades_display.get(prioridad, prioridad))
        c1.border = thin_border
        c2 = ws1.cell(row=row, column=col_offset + 1, value=cantidad)
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='center')
        if prioridad == 'critica' and cantidad > 0:
            c2.fill = rojo_fill
        elif prioridad == 'alta' and cantidad > 0:
            c2.fill = amarillo_fill
        row += 1

    # Por técnico
    row = max(row, 13)
    row += 1
    ws1.cell(row=row, column=1, value="Tickets por Técnico Asignado").font = Font(bold=True, size=11, color='1E3A5F')
    row += 1
    _escribir_fila_header(ws1, row, ['Técnico', 'Tickets'], header_font, header_fill, header_alignment, thin_border)
    row += 1
    for tecnico, cantidad in datos['por_tecnico']:
        ws1.cell(row=row, column=1, value=tecnico).border = thin_border
        c = ws1.cell(row=row, column=2, value=cantidad)
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')
        row += 1

    for col, width in zip('ABCDEF', [25, 12, 5, 20, 12, 5]):
        ws1.column_dimensions[col].width = width

    # === HOJA 2: Detalle ===
    ws2 = wb.create_sheet("Detalle de Tickets")
    ws2.merge_cells('A1:H1')
    ws2['A1'].value = f"Detalle de Tickets - {datos['fecha_inicio']} al {datos['fecha_fin']}"
    ws2['A1'].font = Font(bold=True, size=14, color='1E3A5F')
    ws2['A1'].alignment = Alignment(horizontal='center')

    headers2 = ['Folio', 'Fecha', 'Empleado', 'Categoría', 'Prioridad', 'Estado', 'Técnico', 'Hrs Resolución']
    _escribir_fila_header(ws2, 3, headers2, header_font, header_fill, header_alignment, thin_border)

    for row_idx, ticket in enumerate(datos['tickets'], 4):
        ws2.cell(row=row_idx, column=1, value=ticket.folio).border = thin_border
        ws2.cell(row=row_idx, column=2, value=str(ticket.fecha_creacion.date())).border = thin_border
        ws2.cell(row=row_idx, column=3, value=ticket.empleado.nombre_completo if ticket.empleado else '-').border = thin_border
        ws2.cell(row=row_idx, column=4, value=ticket.get_categoria_display()).border = thin_border
        c_pri = ws2.cell(row=row_idx, column=5, value=ticket.get_prioridad_display() if ticket.prioridad else 'Sin asignar')
        c_pri.border = thin_border
        if ticket.prioridad == 'critica':
            c_pri.fill = rojo_fill
        elif ticket.prioridad == 'alta':
            c_pri.fill = amarillo_fill
        ws2.cell(row=row_idx, column=6, value=ticket.get_estado_display()).border = thin_border
        tecnico = ticket.asignado_a.get_full_name() or ticket.asignado_a.username if ticket.asignado_a else '-'
        ws2.cell(row=row_idx, column=7, value=tecnico).border = thin_border
        c_hrs = ws2.cell(row=row_idx, column=8, value=round(ticket.tiempo_resolucion_horas, 1))
        c_hrs.border = thin_border
        c_hrs.alignment = Alignment(horizontal='center')
        if ticket.estado == 'concluido':
            c_hrs.fill = verde_fill

    for col, width in zip('ABCDEFGH', [14, 12, 28, 14, 12, 14, 22, 16]):
        ws2.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_reporte_permisos_excel(datos):
    """
    Excel de permisos con 2 hojas:
    - Hoja 1: Resumen por estado, tipo, top empleados
    - Hoja 2: Detalle de solicitudes
    """
    wb = openpyxl.Workbook()
    header_font, header_fill, header_alignment, thin_border = _estilos_base()
    aprobado_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    rechazado_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    pendiente_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    estado_fills = {'aprobado': aprobado_fill, 'rechazado': rechazado_fill, 'pendiente': pendiente_fill}

    # === HOJA 1: Resumen ===
    ws1 = wb.active
    ws1.title = "Resumen"

    ws1.merge_cells('A1:F1')
    ws1['A1'].value = f"Reporte de Permisos - {datos['fecha_inicio']} al {datos['fecha_fin']}"
    ws1['A1'].font = Font(bold=True, size=14, color='1E3A5F')
    ws1['A1'].alignment = Alignment(horizontal='center')

    ws1['A3'] = f"Total solicitudes: {datos['total_solicitudes']}"
    ws1['A3'].font = Font(bold=True, size=12)
    ws1['C3'] = f"Días aprobados: {datos['total_dias_aprobados']}"
    ws1['E3'] = f"Pendientes actuales: {len(datos['pendientes_ahora'])}"

    # Por estado
    row = 5
    ws1.cell(row=row, column=1, value="Por Estado").font = Font(bold=True, size=11, color='1E3A5F')
    row += 1
    _escribir_fila_header(ws1, row, ['Estado', 'Cantidad'], header_font, header_fill, header_alignment, thin_border)
    row += 1
    estados_display = {'pendiente': 'Pendiente', 'aprobado': 'Aprobado', 'rechazado': 'Rechazado', 'cancelado': 'Cancelado'}
    for estado, cantidad in datos['por_estado'].items():
        c1 = ws1.cell(row=row, column=1, value=estados_display.get(estado, estado))
        c1.border = thin_border
        c2 = ws1.cell(row=row, column=2, value=cantidad)
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='center')
        if estado in estado_fills and cantidad > 0:
            c2.fill = estado_fills[estado]
        row += 1

    # Por tipo de permiso (columna D)
    row = 5
    ws1.cell(row=row, column=4, value="Por Tipo de Permiso").font = Font(bold=True, size=11, color='1E3A5F')
    row += 1
    _escribir_fila_header(ws1, row, ['Tipo', 'Solicitudes', 'Días'], header_font, header_fill, header_alignment, thin_border)
    row += 1
    for tipo, stats in datos['por_tipo']:
        ws1.cell(row=row, column=4, value=tipo).border = thin_border
        c2 = ws1.cell(row=row, column=5, value=stats['solicitudes'])
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='center')
        c3 = ws1.cell(row=row, column=6, value=stats['dias'])
        c3.border = thin_border
        c3.alignment = Alignment(horizontal='center')
        row += 1

    # Top empleados por días aprobados
    row = max(row, 13) + 1
    ws1.cell(row=row, column=1, value="Top Empleados (días de permiso aprobados)").font = Font(bold=True, size=11, color='1E3A5F')
    row += 1
    _escribir_fila_header(ws1, row, ['Empleado', 'Días Aprobados'], header_font, header_fill, header_alignment, thin_border)
    row += 1
    for nombre, dias in datos['top_empleados']:
        ws1.cell(row=row, column=1, value=nombre).border = thin_border
        c = ws1.cell(row=row, column=2, value=dias)
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')
        row += 1

    for col, width in zip('ABCDEF', [28, 14, 5, 30, 14, 10]):
        ws1.column_dimensions[col].width = width

    # === HOJA 2: Detalle ===
    ws2 = wb.create_sheet("Detalle de Solicitudes")
    ws2.merge_cells('A1:G1')
    ws2['A1'].value = f"Detalle de Solicitudes - {datos['fecha_inicio']} al {datos['fecha_fin']}"
    ws2['A1'].font = Font(bold=True, size=14, color='1E3A5F')
    ws2['A1'].alignment = Alignment(horizontal='center')

    headers2 = ['Empleado', 'Tipo Permiso', 'Fecha Inicio', 'Fecha Fin', 'Días', 'Estado', 'Aprobador']
    _escribir_fila_header(ws2, 3, headers2, header_font, header_fill, header_alignment, thin_border)

    for row_idx, sol in enumerate(datos['solicitudes'], 4):
        ws2.cell(row=row_idx, column=1, value=sol.empleado.nombre_completo if sol.empleado else '-').border = thin_border
        ws2.cell(row=row_idx, column=2, value=str(sol.tipo_permiso)).border = thin_border
        ws2.cell(row=row_idx, column=3, value=str(sol.fecha_inicio)).border = thin_border
        ws2.cell(row=row_idx, column=4, value=str(sol.fecha_fin)).border = thin_border
        c_dias = ws2.cell(row=row_idx, column=5, value=sol.dias_solicitados)
        c_dias.border = thin_border
        c_dias.alignment = Alignment(horizontal='center')
        c_est = ws2.cell(row=row_idx, column=6, value=sol.get_estado_display())
        c_est.border = thin_border
        if sol.estado in estado_fills:
            c_est.fill = estado_fills[sol.estado]
        aprobador = sol.aprobador.nombre_completo if sol.aprobador else '-'
        ws2.cell(row=row_idx, column=7, value=aprobador).border = thin_border

    for col, width in zip('ABCDEFG', [30, 25, 14, 14, 8, 14, 25]):
        ws2.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
